from __future__ import annotations

import csv
import itertools
import operator
import os
import typing

import openpyxl
import openpyxl.styles
import openpyxl.worksheet.worksheet
import pandas as pd

pd.options.plotting.backend = "plotly"
# import matplotlib.pyplot as plt

import model
from model import Category, Class, Range, Participant, Duel, Queue

# RANGE_QUEUES = {
#     model.Range.First: [model.Queue.MODIFIED, model.Queue.STANDARD_MANUAL, ],
#     model.Range.Second: [model.Queue.STANDARD_1, model.Queue.STANDARD_2, model.Queue.LADY, model.Queue.OPEN, ],
# }
#
# # alternate 1
# RANGE_QUEUES = {
#     model.Range.First: [model.Queue.STANDARD_1, model.Queue.LADY, model.Queue.MODIFIED, ],
#     model.Range.Second: [model.Queue.STANDARD_2, model.Queue.OPEN, model.Queue.STANDARD_MANUAL, ],
# }
#
# # alternate 2
# RANGE_QUEUES = {
#     model.Range.First: [model.Queue.STANDARD_1, model.Queue.LADY, model.Queue.OPEN, ],
#     model.Range.Second: [model.Queue.STANDARD_2, model.Queue.MODIFIED, model.Queue.STANDARD_MANUAL, ],
# }

QUEUE_VARIANTS = {
    # "original": {
    #     model.Range.First: [model.Queue.MODIFIED, model.Queue.STANDARD_MANUAL, ],
    #     model.Range.Second: [model.Queue.STANDARD_1, model.Queue.STANDARD_2, model.Queue.LADY, model.Queue.OPEN, ],
    # },
    # "alternate_1": {
    #     model.Range.First: [model.Queue.STANDARD_2, model.Queue.OPEN, model.Queue.MODIFIED, ],
    #     model.Range.Second: [model.Queue.STANDARD_1, model.Queue.LADY,  model.Queue.STANDARD_MANUAL, ],
    # },
    # "single_tactics": {
    #     model.Range.First: [model.Queue.STANDARD_2, model.Queue.OPEN, model.Queue.STANDARD_MANUAL, ],
    #     model.Range.Second: [model.Queue.STANDARD_1, model.Queue.LADY, model.Queue.MODIFIED, ],
    # },
    # "single_tactics_adapted": {
    #     model.Range.First: [model.Queue.STANDARD_2, model.Queue.MODIFIED, model.Queue.STANDARD_MANUAL, ],
    #     model.Range.Second: [model.Queue.STANDARD_1, model.Queue.OPEN, model.Queue.LADY, ],
    # },
    "final": {
        model.Range.First: [
            model.Queue.STANDARD_1,
            model.Queue.MODIFIED,
            model.Queue.OPEN,
        ],
        model.Range.Second: [
            model.Queue.STANDARD_2,
            model.Queue.STANDARD_MANUAL,
            model.Queue.LADY,
        ],
    },
    # "huge_standard": {
    #     model.Range.First: [model.Queue.STANDARD_2, model.Queue.STANDARD_MANUAL, model.Queue.MODIFIED,
    #                         model.Queue.OPEN, ],
    #     model.Range.Second: [model.Queue.STANDARD, model.Queue.LADY, ],
    # },
    # "huge_standard_everyone_once": {
    #     model.Range.First: [model.Queue.STANDARD_2, model.Queue.STANDARD_MANUAL, model.Queue.MODIFIED,
    #                         model.Queue.OPEN, ],
    #     model.Range.Second: [model.Queue.STANDARD, model.Queue.LADY, ],
    # },
}

QUEUE_REPEATS = {
    model.Queue.STANDARD_1: False,
    model.Queue.STANDARD_2: False,
    model.Queue.STANDARD_MANUAL: False,
    model.Queue.MODIFIED: False,
    model.Queue.OPEN: False,
    model.Queue.LADY: True,
}

NONCE = Participant("", "", "", 0, None, None)


def generate_duels(participants: list[Participant], repeat: bool) -> list[Duel]:
    participants = participants[:]
    if len(participants) % 2 != 0:
        participants.append(NONCE)

    top, bottom = (
        participants[: len(participants) // 2],
        participants[len(participants) // 2 :],
    )

    result = []

    for idx in range(len(participants) - 1):
        tb = (top, bottom)
        d = list(zip(top, bottom))
        duels = [Duel(t, b) for t, b in zip(top, bottom)]
        result = result + duels
        top, bottom = rotate(top, bottom)

    if repeat:
        result = result + [d.swapped() for d in result]
    else:
        # for non-repeat tournaments make sure everyone has more or less equal number of duels
        # where they're to the left and to the right
        first_participant = participants[0]
        first_participant_duels = [d for d in result if first_participant in d]
        remaining_duels, swap_duels = (
            first_participant_duels[::2],
            first_participant_duels[1::2],
        )
        for duel in swap_duels:
            idx = result.index(duel)
            result[idx] = duel.swapped()

    result = [d for d in result if NONCE not in d]

    result = fix_disbalanced_pairs(result)
    return result


def fix_disbalanced_pairs(duels: list[Duel]) -> list[Duel]:
    participants = {d.left for d in duels} | {d.right for d in duels}
    participants_leftright = []
    for p in participants:
        count_left = len([d for d in duels if d.left == p])
        count_right = len([d for d in duels if d.right == p])
        participants_leftright.append((count_left, count_right, p))
        print(p.name, count_left, count_right)

    disbalanced_participants = [
        (count_left, count_right, p)
        for count_left, count_right, p in participants_leftright
        if abs(count_left - count_right) > 1
    ]
    if not disbalanced_participants:
        return duels
    result = duels[:]
    disbalanced_participants.sort()
    half_length = len(disbalanced_participants) // 2
    db_left, db_right = (
        disbalanced_participants[:half_length],
        disbalanced_participants[half_length:],
    )
    for left, right in zip(db_left, db_right):
        pleft = left[2]
        pright = right[2]

        old_duel = Duel(pright, pleft)
        new_duel = Duel(pleft, pright)
        idx = result.index(old_duel)
        result[idx] = new_duel

    for p in participants:
        count_left = len([d for d in result if d.left == p])
        count_right = len([d for d in result if d.right == p])
        participants_leftright.append((count_left, count_right, p))
        print(p.name, count_left, count_right)
    return result


def rotate(
    top: list[Participant], bottom: list[Participant]
) -> (list[Participant], list[Participant]):
    current = (top, bottom)
    result = ([top[0]] + [bottom[0]] + top[1 : len(top) - 1], bottom[1:] + [top[-1]])
    return result


def assert_pairs_valid(
    participants: list[Participant], variant_name: str, duels: list[Duel]
) -> bool:
    MIN_DOWNTIME = 1
    MAX_DOWNTIME = 4

    duel_delays = get_participant_delays(duels)

    df = pd.DataFrame(
        [
            {
                "delay_left": delay[0],
                "delay_right": delay[1],
            }
            for delay, duel in duel_delays
        ],
        # {
        #     "delay_left": [delay[0] for delay, duel in duel_delays],
        #     "delay_right": [delay[1] for delay, duel in duel_delays],
        # }
    )
    print(variant_name)
    print(df.describe(include="all"))
    print("======")
    fig = df.plot(
        title=variant_name,
        kind="bar",
        # ylim=(-1, 25),
    )
    fig.show()
    # print(duel_delays)

    return True


def get_participant_delays(duels):
    duel_delays = []
    for idx, duel in enumerate(duels):
        if idx == 0:
            duel_delays.append(((None, None), duel))
            continue
        delay_left = None
        delay_right = None
        for delay, prev_duel in enumerate(reversed(duels[:idx])):
            if delay_left is None and duel.left in prev_duel:
                delay_left = delay
            if delay_right is None and duel.right in prev_duel:
                delay_right = delay
        duel_delays.append(((delay_left, delay_right), duel))
    return duel_delays


def read_participants(path: str) -> list[Participant]:
    result = []
    with open(path) as f:
        reader = csv.DictReader(f, delimiter=";")

        for row in reader:
            for cat in Category:
                if row["category"] == cat.value:
                    row["category"] = cat
                    break
            else:
                raise ValueError(f"No category found for {row}")
            for c in Class:
                if row["clazz"] == c.value:
                    row["clazz"] = c
                    break
            else:
                raise ValueError(f"No class found for {row}")
            p = Participant(**row)
            result.append(p)
    return result


def ensure_ladies_have_guns(std_1: list[Participant], std_2: list[Participant]):
    husbands = ["Рогозін Олександр", "Шатєєв Олексій"]
    for name in husbands:
        for p in std_2:
            if p.name == name:
                p_for_swap = [
                    candidate for candidate in std_1 if candidate.name not in husbands
                ][0]
                std_2.remove(p)
                std_1.remove(p_for_swap)
                std_1.append(p)
                std_2.append(p_for_swap)
                break


def ensure_classes_not_fucked_up(std_1: list[Participant], std_2: list[Participant]):
    swapper_names_1 = [
        "Дубик",
    ]
    swapper_names_2 = [
        "Ситарський",
        "Кольченко",
    ]

    swappers_1 = [p for p in std_1 for name in swapper_names_1 if name in p.name]
    swappers_2 = [p for p in std_2 for name in swapper_names_2 if name in p.name]
    for left, right in zip(swappers_1, swappers_2):
        std_1.remove(left)
        std_2.append(left)

        std_2.remove(right)
        std_1.append(right)


def equalize_std(std_1: list[Participant], std_2: list[Participant]):
    n = "Степаняк Михайло"
    p = [x for x in std_2 if x.name == n][0]
    std_2.remove(p)
    std_1.append(p)


def build_queues(participants: list[Participant]) -> dict[Queue, list[Participant]]:
    men = [p for p in participants if p.category != Category.LADY]
    standard = [p for p in men if p.clazz == Class.STANDARD]

    half = len(standard) // 2
    standard_1, standard_2 = standard[:half], standard[half:]
    #
    # ensure_ladies_have_guns(standard_1, standard_2)
    # ensure_classes_not_fucked_up(standard_1, standard_2)
    # equalize_std(standard_1, standard_2)

    queues = {
        Queue.LADY: [p for p in participants if p.category == Category.LADY],
        Queue.MODIFIED: [p for p in men if p.clazz == Class.MODIFIED],
        Queue.OPEN: [p for p in men if p.clazz == Class.OPEN],
        Queue.STANDARD_MANUAL: [p for p in men if p.clazz == Class.STANDARD_MANUAL],
        Queue.STANDARD_1: standard_1,
        Queue.STANDARD_2: standard_2,
    }

    return queues


def deliver_sorted_pairs(range_queues, excel_writer: pd.ExcelWriter):
    for r, q in range_queues.items():
        # delays = get_participant_delays(q)
        q_items = [
            {
                "number": idx + 1,
                "class": render_class(duel.left),
                "left": duel.left.name,
                "delay_left": None,
                "_": " " * 16,
                "right": duel.right.name,
                "delay_right": None,
                "__": " " * 16,
                # "left_delay": delays[idx][0][0],
                # "right_delay": delays[idx][0][1]
            }
            for idx, duel in enumerate(q)
        ]
        df = pd.DataFrame(q_items)

        df_delays = df[["left", "right"]]
        df_delays["n"] = df_delays.index + 1

        participants = list(pd.concat([df_delays.left, df_delays.right]).unique())
        for p in participants:
            d = df_delays[(df_delays.left == p) | (df_delays.right == p)]
            d = d.join(d.shift(-1), rsuffix="_next")
            d["delay"] = (d.n_next - d.n).astype(int, errors="ignore")
            df["delay_left"].update(d[(d.left == p)].delay)
            df["delay_right"].update(d[(d.right == p)].delay)

        df[["delay_left", "delay_right"]].fillna("F", inplace=True)

        sheet_name = f"Пари Рубіж {r.value} Сортовані"
        header_text = sheet_name
        add_sheet_header(excel_writer, header_text, sheet_name, 6)
        df.to_excel(
            excel_writer,
            sheet_name=sheet_name,
            index=False,
            header=False,
            startrow=1,
        )


def render_class(p: Participant) -> str:
    if p.clazz == Class.STANDARD_MANUAL:
        return "SM"
    elif p.clazz == Class.OPEN:
        return "O"
    elif p.clazz == Class.MODIFIED:
        return "T"
    elif p.clazz == Class.STANDARD:
        return "S" if p.category == Category.GENERAL else "SL"


def render_class_ua(p: Participant) -> str:
    if p.clazz == Class.STANDARD_MANUAL:
        return "Стандарт-Мануал"
    elif p.clazz == Class.OPEN:
        return "Відкритий"
    elif p.clazz == Class.MODIFIED:
        return "Тактика"
    elif p.clazz == Class.STANDARD:
        return "Стандарт" if p.category == Category.GENERAL else "Стандарт Леді"


def deliver_participants(participants: list[Participant], excel_writer: pd.ExcelWriter):
    df = (
        pd.DataFrame(
            [
                {
                    "name": p.name,
                    "class": render_class(p),
                }
                for p in participants
            ]
        )
        .sort_values(
            [
                "class",
                "name",
            ]
        )
        .set_index("class", drop=True)
    )

    df["counts"] = df.groupby(
        [
            "class",
        ]
    ).count()
    df = df.reset_index().set_index(["class", "counts"])
    df = df.rename(
        columns={
            "name": "Імʼя",
            "class": "Клас",
        }
    )
    # cnt = df.groupby(["class", ]).count()
    header_text = "Загальний список"
    sheet_name = header_text
    add_sheet_header(excel_writer, header_text, sheet_name, 3)

    df.to_excel(
        excel_writer,
        sheet_name=sheet_name,
        # index=False,
        merge_cells=True,
        startrow=2,
    )


def deliver_range_lists(
    range: Range,
    participants: typing.Iterable[Participant],
    excel_writer: pd.ExcelWriter,
):
    df = (
        pd.DataFrame(
            [
                {
                    "name": p.name,
                    "class": render_class_ua(p),
                }
                for p in participants
            ]
        )
        .sort_values(["class", "name"])
        .reset_index(drop=True)
    )
    df.index = df.index + 1
    # df = df[["No.", "class", "name"]]

    sheet_name = f"Список Рубіж №{range.value}"
    header_text = f"Рубіж №{range.value}"
    add_sheet_header(excel_writer, header_text, sheet_name, 3)

    df.to_excel(
        excel_writer,
        sheet_name=sheet_name,
        header=False,
        startrow=1,
    )


def deliver_range_pairs(
    range: Range, duels: typing.Iterable[Duel], excel_writer: pd.ExcelWriter
):
    df = (
        pd.DataFrame(
            [
                {
                    "class": render_class(duel.left),
                    "left": duel.left.name,
                    "_": " " * 16,
                    "right": duel.right.name,
                    "__": " " * 16,
                }
                for duel in duels
            ]
        )
        .sort_values(
            [
                "class",
                "left",
                "right",
            ]
        )
        .reset_index(drop=True)
    )
    df.index = df.index + 1

    sheet_name = f"Пари Рубіж №{range.value}"
    header_text = sheet_name
    add_sheet_header(excel_writer, header_text, sheet_name, 6)
    df.to_excel(
        excel_writer,
        sheet_name=sheet_name,
        header=False,
        startrow=1,
    )


def deliver_standard_groups(
    standard_1: list[Participant], standard_2: list[Participant], excel_writer
):
    dataframes = [
        pd.DataFrame({"group": f"Група №{idx+1}", "name": [p.name for p in queue]})
        for idx, queue in enumerate([standard_1, standard_2])
    ]
    df = pd.concat(dataframes)
    df = df.reset_index()
    df.index = df.index + 1
    df = df.set_index(
        [
            "group",
            "index",
        ]
    )

    sheet_name = "Рубіж №2 Групи Стандарт"
    header_text = sheet_name
    add_sheet_header(excel_writer, header_text, sheet_name, 3)

    df.to_excel(excel_writer, sheet_name=sheet_name, merge_cells=True, startrow=2)


def equalize_column_width(excel_writer):
    workbook = excel_writer.book
    for ws in workbook.worksheets:
        worksheet: openpyxl.worksheet.worksheet.Worksheet = ws
        dims = {}
        for row in worksheet.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max(
                        (dims.get(cell.column_letter, 0), len(str(cell.value)))
                    )
        for col, value in dims.items():
            worksheet.column_dimensions[col].width = value + 2


def add_sheet_header(excel_writer, header_text, sheet_name, width: int):
    workbook: openpyxl.Workbook = excel_writer.book
    worksheet: openpyxl.worksheet.worksheet.Worksheet = workbook.create_sheet(
        sheet_name
    )
    excel_writer.sheets[sheet_name] = worksheet
    header_font = openpyxl.styles.Font(sz=14, bold=True)
    header_cell = worksheet["A1"]
    worksheet["A1"] = header_text
    worksheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=width)
    header_cell.font = header_font
    header_cell.alignment = openpyxl.styles.Alignment(horizontal="center")


def merge_queues(queues: list[list[Duel]]) -> list[Duel]:
    percentages = [[(duel, idx / len(q)) for idx, duel in enumerate(q)] for q in queues]
    sorted_duels = sorted(
        itertools.chain.from_iterable(percentages), key=operator.itemgetter(1)
    )
    return [duel for duel, percentage in sorted_duels]


def main():
    participants = read_participants("../../participants.csv")
    queues = build_queues(participants)

    for variant_name, variant in QUEUE_VARIANTS.items():
        queue_duels = {
            queue_name: generate_duels(queue_participants, QUEUE_REPEATS[queue_name])
            for queue_name, queue_participants in queues.items()
        }

        range_duels = {
            range: merge_queues([queue_duels[queue_name] for queue_name in queue_names])
            for range, queue_names in variant.items()
        }

        try:
            os.mkdir("../../out")
        except FileExistsError:
            pass
        target_dir = f"out/{variant_name}"
        try:
            os.mkdir(target_dir)
        except FileExistsError:
            pass
        excel_writer = pd.ExcelWriter(f"{target_dir}/pairs_{variant_name}.xlsx")

        deliver_participants(participants, excel_writer)

        for range, duels in range_duels.items():
            range_participants = set(itertools.chain(*duels))
            deliver_range_lists(range, range_participants, excel_writer)

        deliver_standard_groups(
            queues[Queue.STANDARD_1], queues[Queue.STANDARD_2], excel_writer
        )

        for range, duels in range_duels.items():
            deliver_range_pairs(range, duels, excel_writer)

        deliver_sorted_pairs(range_duels, excel_writer)
        equalize_column_width(excel_writer)

        excel_writer.save()


# Press the green button in the gutter to run the script.

if __name__ == "__main__":
    main()
# See PyCharm help at https://www.jetbrains.com/help/pycharm/
