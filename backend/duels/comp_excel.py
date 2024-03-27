from __future__ import annotations

import itertools
import pickle
import typing

import openpyxl
import pandas as pd
import xlsxwriter
import xlsxwriter.worksheet

from duels.model import Participant, Class, Range, Duel


ExcelInputType = typing.TypeVar("ExcelInputType", bound=dict[Range, list[Duel]])


def deliver_excel(duels: ExcelInputType, path: str):
    writer = pd.ExcelWriter(path)
    range_participants = _get_participants_per_range(duels)
    all_participants = itertools.chain(*range_participants.values())

    _deliver_participants(all_participants, writer)
    ranges = [Range(r) for r in range_participants.keys()]
    for rng in ranges:
        _deliver_range_lists(rng, range_participants[rng], writer)

    _deliver_all_groups(range_participants, writer)

    deliver_sorted_pairs(duels, writer)
    for rng in ranges:
        _deliver_range_results(rng, duels[rng], range_participants[rng], writer)

    for clazz in Class.__members__.values():
        deliver_blank_sheet(clazz, writer.book)

    writer.close()


def _get_participants_per_range(
    duels: dict[Range, list[Duel]]
) -> dict[str, list[Participant]]:
    result = {}

    for rng in duels:
        participants = [d.left for d in duels[rng]] + [d.right for d in duels[rng]]
        participants = list(sorted(set(participants)))
        result[rng] = participants

    return result


def _deliver_participants(
    participants: typing.Iterable[Participant], excel_writer: pd.ExcelWriter
):
    df = (
        pd.DataFrame(
            [
                {
                    "name": p.name,
                    "class": _render_class(p),
                }
                for p in participants
            ]
        )
        .sort_values(
            [
                "class",
                "name",
            ]
        )
        .set_index("class", drop=True)
    )

    df["counts"] = df.groupby(
        [
            "class",
        ]
    ).count()
    df = df.reset_index().set_index(["class", "counts"])
    df = df.rename(
        columns={
            "name": "Імʼя",
            "class": "Клас",
        }
    )
    # cnt = df.groupby(["class", ]).count()
    header_text = "Загальний список"
    sheet_name = header_text
    _add_sheet_header(excel_writer, header_text, sheet_name, 3)

    df.to_excel(
        excel_writer,
        sheet_name=sheet_name,
        # index=False,
        merge_cells=True,
        startrow=2,
    )
    worksheet: xlsxwriter.worksheet.Worksheet = excel_writer.sheets[sheet_name]
    worksheet.autofit()


def _deliver_range_lists(
    rng: Range,
    participants: typing.Iterable[Participant],
    excel_writer: pd.ExcelWriter,
):
    df = (
        pd.DataFrame(
            [
                {
                    "name": p.name,
                    "class": _render_class(p),
                }
                for p in participants
            ],
            columns=['class', 'name'],
        )
        .sort_values(["class", "name"])
        .reset_index(drop=True)
    )
    df.index = df.index + 1
    # df = df[["No.", "class", "name"]]

    sheet_name = _sheet_name_range_list(rng)
    header_text = f"Рубіж №{rng.value}"
    _add_sheet_header(excel_writer, header_text, sheet_name, 3)

    df.to_excel(
        excel_writer,
        sheet_name=sheet_name,
        header=False,
        startrow=1,
    )
    worksheet: xlsxwriter.worksheet.Worksheet = excel_writer.sheets[sheet_name]
    worksheet.autofit()


def _deliver_range_pairs(
    range: Range, duels: typing.Iterable[Duel], excel_writer: pd.ExcelWriter
):

    df = (
        pd.DataFrame(
            [
                {
                    "class": _render_class(duel.left),
                    "left": duel.left.name,
                    "_": " " * 16,
                    "right": duel.right.name,
                    "__": " " * 16,
                }
                for duel in duels
            ]
        )
        .sort_values(
            [
                "class",
                "left",
                "right",
            ]
        )
        .reset_index(drop=True)
    )
    df.index = df.index + 1

    sheet_name = f"Пари Рубіж №{range.value}"
    header_text = sheet_name
    _add_sheet_header(excel_writer, header_text, sheet_name, 6)
    df.to_excel(
        excel_writer,
        sheet_name=sheet_name,
        header=False,
        startrow=1,
    )
    worksheet: xlsxwriter.worksheet.Worksheet = excel_writer.sheets[sheet_name]
    worksheet.autofit()


def _deliver_all_groups(
    range_participants: dict[str, list[Participant]], excel_writer: pd.ExcelWriter
):
    range_classes = {
        rng: {p.clazz for p in participants}
        for rng, participants in range_participants.items()
    }
    repeated_classes = [
        c
        for c in Class.__members__.values()
        if all(c in range_class for range_class in range_classes.values())
    ]
    for clazz in repeated_classes:
        groups = {
            rng: [p for p in participants if p.clazz == clazz]
            for rng, participants in range_participants.items()
        }
        _deliver_groups(groups.get(Range.First, {}), groups.get(Range.Second, {}), excel_writer)


def _deliver_groups(
    group_1: list[Participant], group_2: list[Participant], excel_writer
):
    dataframes = [
        pd.DataFrame({"group": f"Група №{idx + 1}", "name": [p.name for p in queue]})
        for idx, queue in enumerate([group_1, group_2])
    ]
    df = pd.concat(dataframes)
    df = df.reset_index()
    df.index = df.index + 1
    df = df.set_index(
        [
            "group",
            "index",
        ]
    )

    sheet_name = f"Групи {_render_class_ua(group_1[0])}"
    header_text = sheet_name
    _add_sheet_header(excel_writer, header_text, sheet_name, 3)

    df.to_excel(excel_writer, sheet_name=sheet_name, merge_cells=True, startrow=2)
    worksheet: xlsxwriter.worksheet.Worksheet = excel_writer.sheets[sheet_name]
    worksheet.autofit()


def _render_class(p: Participant) -> str:
    if p.clazz == Class.STANDARD_MANUAL:
        return "SM"
    elif p.clazz == Class.OPEN:
        return "O"
    elif p.clazz == Class.MODIFIED:
        return "T"
    elif p.clazz == Class.STANDARD:
        return "S"
    elif p.clazz == Class.STANDARD_LADY:
        return "SL"


def _render_class_ua(p: Participant | Class) -> str:
    if isinstance(p, Participant):
        p = p.clazz
    if p == Class.STANDARD_MANUAL:
        return "Стандарт-Мануал"
    elif p == Class.OPEN:
        return "Відкритий"
    elif p == Class.MODIFIED:
        return "Тактика"
    elif p == Class.STANDARD:
        return "Стандарт"
    elif p == Class.STANDARD_LADY:
        return "Стандарт Леді"


def _add_sheet_header(excel_writer, header_text, sheet_name, width: int):
    workbook: xlsxwriter.Workbook = excel_writer.book
    worksheet: xlsxwriter.worksheet.Worksheet = workbook.add_worksheet(sheet_name)
    excel_writer.sheets[sheet_name] = worksheet
    header_format = workbook.add_format(
        {
            "bold": True,
            "align": "center",
            "font_size": 16,
            "bottom": 1,
        }
    )
    worksheet.merge_range(0, 0, 0, width - 1, header_text, header_format)
    # header_cell.font = header_font
    # header_cell.alignment = openpyxl.styles.Alignment(horizontal="center")


def _add_sheet_header_openpyxl(excel_writer, header_text, sheet_name, width: int):
    workbook: openpyxl.Workbook = excel_writer.book
    worksheet: openpyxl.worksheet.worksheet.Worksheet = workbook.create_sheet(
        sheet_name
    )
    excel_writer.sheets[sheet_name] = worksheet
    header_font = openpyxl.styles.Font(sz=14, bold=True)
    header_cell = worksheet["A1"]
    worksheet["A1"] = header_text
    worksheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=width)
    header_cell.font = header_font
    header_cell.alignment = openpyxl.styles.Alignment(horizontal="center")


def _equalize_column_width_openpyxl(excel_writer):
    workbook = excel_writer.book
    for ws in workbook.worksheets:
        worksheet: openpyxl.worksheet.worksheet.Worksheet = ws
        dims = {}
        for row in worksheet.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max(
                        (dims.get(cell.column_letter, 0), len(str(cell.value)))
                    )
        for col, value in dims.items():
            worksheet.column_dimensions[col].width = value + 2


def deliver_blank_sheet(clazz: Class, workbook: xlsxwriter.Workbook):
    fmt = workbook.add_format(
        {
            "bottom": 1,
            "right": 1,
        }
    )
    fmt_bold = workbook.add_format(
        {
            "bottom": 1,
            "right": 5,
        }
    )
    fmt_header = workbook.add_format(
        {
            "bold": True,
            "align": "center",
            "font_size": 16,
            "bottom": 1,
        }
    )
    sheets = [
        (
            "Вихід з групи",
            [
                ("Вихід з групи", 25),
            ],
        ),
        (
            "Фінали",
            [
                ("Півфінал", 15),
                ("Третє місце", 6),
                ("Фінал", 6),
            ],
        ),
    ]
    header_width = 5

    for sheet_title, subheaders in sheets:
        sheet = workbook.add_worksheet(f"{_render_class_ua(clazz)} {sheet_title}")
        current_row = 0
        for subheader, row_count in subheaders:
            sheet.merge_range(
                current_row,
                0,
                current_row,
                header_width - 1,
                f"{_render_class_ua(clazz)} — {subheader}",
                fmt_header,
            )
            current_row += 1

            for row in range(row_count):
                table_row = current_row + row
                sheet.write_number(table_row, 0, row + 1, fmt_bold)
                sheet.write_string(table_row, 1, " " * 64, fmt)
                sheet.write_string(table_row, 2, " " * 12, fmt_bold)
                sheet.write_string(table_row, 3, " " * 64, fmt)
                sheet.write_string(table_row, 4, " " * 12, fmt_bold)

            current_row += row_count
        sheet.autofit()
        sheet.set_column("B:B", 30, fmt)
        sheet.set_column("D:D", 30, fmt)
        sheet.set_column("A:A", 4, fmt)


def deliver_sorted_pairs(
    range_queues: dict[Range, list[Duel]], excel_writer: pd.ExcelWriter
):
    fmt = excel_writer.book.add_format(
        {
            "bottom": 1,
        }
    )
    fmt_right_border_bold = excel_writer.book.add_format(
        {
            "bottom": 1,
            "right": 5,
        }
    )
    fmt_right_border = excel_writer.book.add_format(
        {
            "bottom": 1,
            "right": 1,
        }
    )
    for r, q in range_queues.items():
        # delays = get_participant_delays(q)
        q_items = [
            {
                "number": idx + 1,
                "class": _render_class(duel.left),
                "left": duel.left.name,
                "_": " " * 16,
                "right": duel.right.name,
            }
            for idx, duel in enumerate(q)
        ]
        df = pd.DataFrame(q_items)

        sheet_name = _sheet_name_range_pairs(r)
        header_text = sheet_name
        _add_sheet_header(excel_writer, header_text, sheet_name, 6)
        df.to_excel(
            excel_writer,
            sheet_name=sheet_name,
            index=False,
            header=False,
            startrow=1,
        )
        worksheet: xlsxwriter.worksheet.Worksheet = excel_writer.sheets[sheet_name]
        worksheet.set_column(
            "A:E",
            1,
            fmt,
        )
        worksheet.set_column("C:C", 1, fmt_right_border)
        worksheet.set_column("E:E", 1, fmt_right_border)
        worksheet.autofit()
        worksheet.set_column("D:D", 10, fmt_right_border_bold)
        worksheet.set_column("F:F", 10, fmt_right_border_bold)
        worksheet.set_column("A:A", 4, fmt_right_border_bold)


def _deliver_range_results(
    r: Range, queue: list[Duel], participants: list[Participant], writer: pd.ExcelWriter
):
    book: xlsxwriter.Workbook = writer.book

    w = ExcelWriter(queue, participants, book)
    w.render_result_sheet(r)


class ExcelWriter:
    duels: list[Duel]
    participants: list[Participant]
    book: xlsxwriter.Workbook

    def __init__(
        self,
        duels: list[Duel],
        participants: list[Participant],
        book: xlsxwriter.Workbook,
    ):
        self.duels = duels
        self.participants = participants
        self.book = book
        self.fmt_header = self.book.add_format(
            {
                "bold": 1,
                "align": "center",
                "bottom": 1,
                "text_wrap": True,
            }
        )
        self.fmt_class_header = self.book.add_format(
            {
                "bold": 1,
                "align": "center",
                "valign": "vcenter",
                "top": 1,
            }
        )
        self.fmt_top_in_class = self.book.add_format(
            {"bg_color": "#C6EFCE", "font_color": "#006100"}
        )
        self.fmt_victory_record = self.book.add_format(
            {
                "bold": 1,
                "align": "center",
            }
        )
        self.fmt_victory = self.book.add_format(
            {
                "bg_color": "#C6EFCE",
                "bold": 1,
                "align": "center",
            }
        )

    def render_result_sheet(self, rng: Range):
        sheet = self.book.add_worksheet(_sheet_name_range_results(rng))
        self._write_duels(sheet)
        self._write_participants(sheet)
        sheet.autofit()


    def _render_victory(self, duel: Duel, left: bool) -> typing.Union[str, int]:
        if not duel.victories:
            return " " * 16
        left_won, right_won = duel.victories
        if left and left_won:
            return 1
        if not left and right_won:
            return 1
        return 0

    def _write_duels(self, sheet: xlsxwriter.worksheet.Worksheet):
        sheet.write(
            0, self._col_duel_number(), "#", self.fmt_header,
        )
        sheet.write(
            0, self._col_duel_participant(True), "Стрілець зліва", self.fmt_header
        )
        sheet.write(0, self._col_duel_result(True), "Перемога\nзліва", self.fmt_header)
        sheet.write(
            0, self._col_duel_result(False), "Перемога\nсправа", self.fmt_header
        )
        sheet.write(
            0, self._col_duel_participant(False), "Стрілець справа", self.fmt_header
        )

        row_first, row_last = self._rows_duels()
        rows = range(row_first, row_last)
        for duel_number, (row, duel) in enumerate(zip(rows, self.duels), 1):
            sheet.write(
                row,
                self._col_duel_number(),
                duel_number
            )
            sheet.write(
                row,
                self._col_duel_participant(True),
                self._render_participant(duel.left),
            )
            sheet.write(
                row,
                self._col_duel_result(True),
                self._render_victory(duel, True)
            )

            sheet.write(
                row,
                self._col_duel_participant(False),
                self._render_participant(duel.right),
            )
            sheet.write(
                row,
                self._col_duel_result(False),
                self._render_victory(duel, False)
            )


        validation_rule = {
            "validate": "list",
            "source": [0, 1],
        }
        sheet.data_validation(
            row_first,
            self._col_duel_result(True),
            row_last,
            self._col_duel_result(False),
            validation_rule,
        )
        sheet.conditional_format(
            row_first,
            self._col_duel_result(True),
            row_last,
            self._col_duel_result(False),
            {
                "type": "cell",
                "criteria": "equal to",
                "value": 1,
                "format": self.fmt_victory,
            },
        )

    def _write_participants(self, sheet: xlsxwriter.worksheet.Worksheet):
        sheet.write(
            0,
            self._col_participant_class(),
            "Клас",
            self.fmt_header,
        )
        sheet.write(
            0,
            self._col_participant(),
            "Стрілець",
            self.fmt_header,
        )
        sheet.write(
            0,
            self._col_victory_count(),
            "Перемог",
            self.fmt_header,
        )
        row_start, row_end = self._rows_participants()
        rows = range(row_start, row_end)
        participants = sorted(self.participants, key=lambda p: (p.clazz, p.name))
        fmt_class_start = self.book.add_format(
            {
                "top": 1,
            }
        )
        for row, p in zip(rows, participants):
            fmt = None
            class_start_row, _ = self._rows_participant_class(p.clazz)
            if row == class_start_row:
                fmt = fmt_class_start
            sheet.write(row, self._col_participant(), self._render_participant(p), fmt)
            address = f"INDIRECT(ADDRESS({row+1}, {self._col_participant()+1}))"
            sheet.write(
                row,
                self._col_victory_count(),
                f"""=SUMIF(B:B, {address}, C:C) + SUMIF(E:E, {address}, D:D)""",
                fmt,
            )

        for clazz in self._classes():
            row_start, row_end = self._rows_participant_class(clazz)
            row_end -= 1
            sheet.merge_range(
                row_start,
                self._col_participant_class(),
                row_end,
                self._col_participant_class(),
                clazz,
                self.fmt_class_header,
            )
            sheet.conditional_format(
                row_start,
                self._col_victory_count(),
                row_end,
                self._col_victory_count(),
                {
                    "type": "top",
                    "value": 4,
                    "format": self.fmt_top_in_class,
                },
            )

    def _render_participant(self, p: Participant) -> str:
        return f"{p.name} {_render_class(p)}"

    def _participant_count(self) -> int:
        return len(self.participants)

    def _duel_count(self) -> int:
        return len(self.duels)

    def _classes(self) -> typing.Iterable[Class]:
        return sorted({p.clazz for p in self.participants})

    def _rows_duels(self) -> (int, int):
        return 1, self._duel_count() + 1

    def _rows_participants(self) -> (int, int):
        return 1, self._participant_count() + 1

    def _col_duel_number(self) -> int:
        return 0

    def _col_duel_participant(self, left: bool) -> int:
        if left:
            return 1
        return 4

    def _col_duel_result(self, left: bool) -> int:
        if left:
            return 2
        return 3

    def _col_participant_class(self) -> int:
        return 7

    def _col_participant(self) -> int:
        return 8

    def _col_victory_count(self) -> int:
        return 9

    def _rows_participant_class(self, clazz: Class) -> (int, int):
        start = None
        for row, p in enumerate(
            sorted(self.participants, key=lambda p: (p.clazz, p.name))
        ):
            row = row + 1  # account for header
            if not start and p.clazz == clazz:
                start = row
            if start and p.clazz != clazz:
                return start, row
        return start, len(self.participants) + 1


def _sheet_name_range_pairs(r):
    return f"Пари — Рубіж №{r.value}"


def _sheet_name_range_results(r):
    return f"Пари Рубіж {r.value} Відомість"


def _sheet_name_range_list(rng):
    return f"Список Рубіж №{rng.value}"


def main():
    with open("duels.pickle", "rb") as f:
        duels = pickle.load(f)
    path = "out.xlsx"
    deliver_excel(duels, path)
    from subprocess import call

    call(["open", path])


if __name__ == "__main__":
    main()
